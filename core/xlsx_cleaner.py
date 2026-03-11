"""
Обработка .xlsx файлов: замена текста в ячейках Excel.

Стратегия:
1. Открываем рабочую книгу через openpyxl
2. Проходим по всем листам и ячейкам
3. Ищем совпадения по regex-паттернам
4. Заменяем текст, сохраняя формулы нетронутыми
"""

import re
import logging
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger('CompanyCleaner')


def is_openpyxl_available() -> bool:
    """Проверяет, установлен ли openpyxl."""
    return HAS_OPENPYXL


def _get_replacement_text(rule: dict, matched_text: str) -> str:
    """Получает текст замены из правила."""
    if "mapper" in rule and rule["mapper"] is not None:
        return rule["mapper"].get_replacement(matched_text)
    return rule.get("replacement", "[ЗАГЛУШКА]")


def clean_xlsx(
    filepath: str,
    output_path: str,
    replacement_rules: list[dict],
) -> dict:
    """
    Заменяет вхождения на заглушки в .xlsx файле.

    replacement_rules: список словарей с ключами:
      - patterns: list[re.Pattern]
      - replacement: str ИЛИ mapper: ReplacementMapper
      - type: str (для статистики)
    """
    if not HAS_OPENPYXL:
        return {
            "status": "error",
            "matches": {},
            "total_replacements": 0,
            "error_message": "Библиотека openpyxl не установлена. "
                             "Установите: pip install openpyxl",
        }

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        logger.error(f"Не удалось открыть {filepath}: {e}")
        return {
            "status": "error",
            "matches": {},
            "total_replacements": 0,
            "error_message": str(e),
        }

    stats = {}  # type -> count

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                # Пропускаем формулы
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    continue
                # Работаем только с текстовыми ячейками
                if not isinstance(cell.value, str):
                    continue

                original = cell.value
                new_text = original

                # Собираем замены
                replacements = []
                for rule in replacement_rules:
                    rule_type = rule.get("type", "custom")
                    for pattern in rule.get("patterns", []):
                        for match in pattern.finditer(new_text):
                            repl = _get_replacement_text(rule, match.group())
                            replacements.append(
                                (match.start(), match.end(), repl, rule_type)
                            )

                if not replacements:
                    continue

                # Убираем пересекающиеся (приоритет длинным)
                replacements.sort(key=lambda r: (r[0], -(r[1] - r[0])))
                filtered = []
                last_end = 0
                for start, end, repl, rtype in replacements:
                    if start >= last_end:
                        filtered.append((start, end, repl, rtype))
                        last_end = end

                if not filtered:
                    continue

                # Подсчёт
                for _, _, _, rtype in filtered:
                    stats[rtype] = stats.get(rtype, 0) + 1

                # Строим новый текст
                result = ''
                pos = 0
                for start, end, repl, _ in filtered:
                    result += new_text[pos:start]
                    result += repl
                    pos = end
                result += new_text[pos:]

                cell.value = result

    try:
        wb.save(output_path)
    except Exception as e:
        logger.error(f"Не удалось сохранить {output_path}: {e}")
        return {
            "status": "error",
            "matches": stats,
            "total_replacements": sum(stats.values()),
            "error_message": str(e),
        }

    return {
        "status": "success",
        "matches": stats,
        "total_replacements": sum(stats.values()),
        "error_message": None,
    }


def preview_xlsx(
    filepath: str,
    replacement_rules: list[dict],
    context_chars: int = 30,
) -> dict:
    """
    Сканирует .xlsx файл БЕЗ изменений, возвращает найденные вхождения.
    """
    if not HAS_OPENPYXL:
        return {
            "status": "error",
            "error_message": "openpyxl не установлен",
            "matches": [],
        }

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return {"status": "error", "error_message": str(e), "matches": []}

    matches = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                if cell.value.startswith('='):
                    continue

                text = cell.value
                text_matches = []

                for rule in replacement_rules:
                    rule_type = rule.get("type", "custom")
                    for pattern in rule.get("patterns", []):
                        for match in pattern.finditer(text):
                            repl = _get_replacement_text(rule, match.group())
                            text_matches.append(
                                (match.start(), match.end(), match.group(),
                                 repl, rule_type)
                            )

                text_matches.sort(key=lambda m: (m[0], -(m[1] - m[0])))
                last_end = 0
                for mstart, mend, original, repl, rule_type in text_matches:
                    if mstart < last_end:
                        continue
                    last_end = mend
                    ctx_start = max(0, mstart - context_chars)
                    ctx_end = min(len(text), mend + context_chars)
                    context = text[ctx_start:ctx_end]
                    if ctx_start > 0:
                        context = '...' + context
                    if ctx_end < len(text):
                        context = context + '...'

                    cell_ref = f"{ws.title}!{cell.coordinate}"
                    matches.append({
                        "original": original,
                        "replacement": repl,
                        "context": context,
                        "type": rule_type,
                        "cell": cell_ref,
                    })

    wb.close()

    type_counts = {}
    for m in matches:
        t = m["type"]
        type_counts[t] = type_counts.get(t, 0) + 1

    return {
        "status": "success",
        "matches": matches,
        "type_counts": type_counts,
    }


def extract_text_xlsx(filepath: str) -> str:
    """Извлекает весь текст из .xlsx файла для автодетекции."""
    if not HAS_OPENPYXL:
        return ""

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return ""

    parts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if not cell.value.startswith('='):
                        parts.append(cell.value)

    wb.close()
    return '\n'.join(parts)
